"""
Excelエクスポートモジュール
議事録・書き起こしのExcel形式出力
"""

import logging
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from export.common import ExportOptions, atomic_save
from time_utils import format_time_hms

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel形式エクスポーター（.xlsx）"""

    def __init__(self):
        self.has_openpyxl = self._check_openpyxl()

    def _check_openpyxl(self) -> bool:
        """openpyxlがインストールされているかチェック"""
        try:
            import openpyxl

            return True
        except ImportError:
            logger.warning("openpyxl not installed. Excel export unavailable.")
            return False

    def export_transcription(
        self,
        segments: List[Dict],
        output_path: str,
        options: Optional[ExportOptions] = None,
    ) -> bool:
        """
        書き起こしをExcel形式でエクスポート

        Args:
            segments: 書き起こしセグメント
            output_path: 出力ファイルパス
            options: エクスポートオプション

        Returns:
            成功したかどうか
        """
        if not self.has_openpyxl:
            logger.error("openpyxl is required for Excel export")
            return False

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "書き起こし"

            # ヘッダー行
            headers = ["No.", "開始時間", "終了時間", "話者", "テキスト"]
            ws.append(headers)

            # ヘッダースタイル
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # データ行
            for i, segment in enumerate(segments, 1):
                start_time = self._format_time(segment.get("start", 0))
                end_time = self._format_time(segment.get("end", 0))
                speaker = segment.get("speaker", "Unknown")
                text = segment.get("text", "")
                ws.append([i, start_time, end_time, speaker, text])

            # 列幅調整
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 60

            # 枠線
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
                for cell in row:
                    cell.border = thin_border

            # 自動折り返し
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            with atomic_save(output_path) as tmp_path:
                wb.save(tmp_path)
            logger.info(f"Excel transcription exported to {output_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export Excel transcription: {e}", exc_info=True)
            return False

    def export_meeting_minutes(  # noqa: C901
        self,
        minutes_data: Dict,
        output_path: str,
        options: Optional[ExportOptions] = None,
    ) -> bool:
        """
        議事録をExcel形式でエクスポート（AGEC社内テンプレート対応）

        Args:
            minutes_data: 議事録データ
            output_path: 出力ファイルパス
            options: エクスポートオプション

        Returns:
            成功したかどうか
        """
        if not self.has_openpyxl:
            logger.error("openpyxl is required for Excel export")
            return False

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "議事録"

            options = options or ExportOptions()
            company = options.company_name

            # タイトル
            ws.merge_cells("A1:E1")
            title_cell = ws["A1"]
            title_cell.value = f"{company} 議事録"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # 基本情報
            row = 3
            ws[f"A{row}"] = "会議名"
            ws[f"B{row}"] = minutes_data.get("title", "")
            ws.merge_cells(f"B{row}:E{row}")
            row += 1

            ws[f"A{row}"] = "日時"
            ws[f"B{row}"] = minutes_data.get("date", "")
            ws[f"C{row}"] = "場所"
            ws[f"D{row}"] = minutes_data.get("location", "")
            ws.merge_cells(f"D{row}:E{row}")
            row += 1

            # 出席者
            ws[f"A{row}"] = "出席者"
            attendees = minutes_data.get("attendees", [])
            ws[f"B{row}"] = ", ".join(attendees) if attendees else ""
            ws.merge_cells(f"B{row}:E{row}")
            row += 2

            # 議題
            if minutes_data.get("agenda"):
                ws[f"A{row}"] = "議題"
                ws[f"A{row}"].font = Font(bold=True)
                row += 1
                for agenda_item in minutes_data["agenda"]:
                    ws[f"B{row}"] = f"• {agenda_item}"
                    ws.merge_cells(f"B{row}:E{row}")
                    row += 1
                row += 1

            # 決定事項
            if minutes_data.get("decisions"):
                ws[f"A{row}"] = "決定事項"
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"A{row}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                row += 1
                for i, decision in enumerate(minutes_data["decisions"], 1):
                    ws[f"A{row}"] = f"{i}."
                    ws[f"B{row}"] = decision
                    ws.merge_cells(f"B{row}:E{row}")
                    row += 1
                row += 1

            # 確認事項
            if minutes_data.get("confirmations"):
                ws[f"A{row}"] = "確認事項"
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"A{row}"].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                row += 1
                for i, confirmation in enumerate(minutes_data["confirmations"], 1):
                    ws[f"A{row}"] = f"{i}."
                    ws[f"B{row}"] = confirmation
                    ws.merge_cells(f"B{row}:E{row}")
                    row += 1
                row += 1

            # アクションアイテム
            if minutes_data.get("action_items"):
                ws[f"A{row}"] = "アクションアイテム"
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"A{row}"].fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                ws[f"B{row}"] = "内容"
                ws[f"C{row}"] = "担当"
                ws[f"D{row}"] = "期限"
                ws[f"E{row}"] = "優先度"
                for cell in ws[row]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                row += 1

                for item in minutes_data["action_items"]:
                    ws[f"A{row}"] = "☐"
                    ws[f"B{row}"] = item.get("description", "")
                    ws[f"C{row}"] = item.get("assignee", "")
                    ws[f"D{row}"] = item.get("due_date", "")
                    ws[f"E{row}"] = item.get("priority", "中")
                    row += 1
                row += 1

            # 議事内容
            if minutes_data.get("statements"):
                ws[f"A{row}"] = "議事内容"
                ws[f"A{row}"].font = Font(bold=True)
                row += 1

                current_speaker = None
                for stmt in minutes_data["statements"]:
                    speaker = stmt.get("speaker", "Unknown")
                    text = stmt.get("text", "")
                    stmt_type = stmt.get("statement_type", "")

                    if speaker != current_speaker:
                        ws[f"A{row}"] = speaker
                        ws[f"A{row}"].font = Font(bold=True, color="4472C4")
                        current_speaker = speaker
                    else:
                        ws[f"A{row}"] = ""

                    prefix = ""
                    if stmt_type == "決定事項":
                        prefix = "【決定】"
                    elif stmt_type == "アクションアイテム":
                        prefix = "【TODO】"
                    elif stmt_type == "確認事項":
                        prefix = "【確認】"

                    ws[f"B{row}"] = f"{prefix}{text}"
                    ws.merge_cells(f"B{row}:E{row}")
                    row += 1

            # 次回会議
            if minutes_data.get("next_meeting"):
                row += 1
                ws[f"A{row}"] = "次回会議"
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"] = minutes_data["next_meeting"]
                ws.merge_cells(f"B{row}:E{row}")

            # 列幅調整
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 12

            # 枠線
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            for row_cells in ws.iter_rows():
                for cell in row_cells:
                    if cell.value:
                        cell.border = thin_border

            with atomic_save(output_path) as tmp_path:
                wb.save(tmp_path)
            logger.info(f"Excel meeting minutes exported to {output_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export Excel meeting minutes: {e}", exc_info=True)
            return False

    def _format_time(self, seconds: float) -> str:
        """秒数を時:分:秒形式に変換"""
        return str(format_time_hms(seconds))


# グローバルインスタンス
_excel_exporter = None
_excel_exporter_lock = threading.Lock()


def get_excel_exporter() -> ExcelExporter:
    """Excelエクスポーターのシングルトンインスタンスを取得"""
    global _excel_exporter
    if _excel_exporter is None:
        with _excel_exporter_lock:
            if _excel_exporter is None:
                _excel_exporter = ExcelExporter()
    return _excel_exporter
